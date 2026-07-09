import os
import gspread
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.service_account import Credentials

SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID")
GOOGLE_CREDENTIALS_JSON = os.environ.get("GOOGLE_CREDENTIALS_JSON")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

WARNA_KATEGORI = {
    "Makanan":       "FF6B6B",
    "Minuman":       "4ECDC4",
    "Transport":     "45B7D1",
    "Kuota/Internet":"96CEB4",
    "Kesehatan":     "FFEAA7",
    "Belanja":       "DDA0DD",
    "Hiburan":       "98D8C8",
    "Tagihan":       "F7DC6F",
    "Lainnya":       "BDC3C7",
}


def get_data_dari_sheets(bulan_str: str) -> list:
    """Ambil data dari sheet bulan tertentu (format: '2025-01')"""
    creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    gc = gspread.authorize(creds)
    
    try:
        spreadsheet = gc.open_by_key(SPREADSHEET_ID)
        sheet = spreadsheet.worksheet(bulan_str)
        return sheet.get_all_records()
    except Exception:
        return []


def buat_excel_dari_data(data: list, nama_file: str, label_bulan: str) -> str | None:
    if not data:
        return None

    wb = Workbook()
    
    # ── Sheet 1: REKAP PER KATEGORI ──────────────────────────────────────────
    ws_rekap = wb.active
    ws_rekap.title = "Rekap Kategori"

    # Kelompokkan per kategori
    per_kategori = {}
    for r in data:
        kat = r.get("Kategori", "Lainnya")
        if kat not in per_kategori:
            per_kategori[kat] = []
        per_kategori[kat].append(r)

    total_keseluruhan = sum(r["Harga"] for r in data)

    # Header
    ws_rekap.merge_cells("A1:E1")
    judul = ws_rekap["A1"]
    judul.value = f"REKAP PENGELUARAN {label_bulan.upper()}"
    judul.font = Font(bold=True, size=14, color="FFFFFF")
    judul.fill = PatternFill("solid", fgColor="2C3E50")
    judul.alignment = Alignment(horizontal="center", vertical="center")
    ws_rekap.row_dimensions[1].height = 30

    headers = ["Kategori", "Jumlah Transaksi", "Total (Rp)", "% dari Total", ""]
    for col, h in enumerate(headers, 1):
        cell = ws_rekap.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="34495E")
        cell.alignment = Alignment(horizontal="center")

    row = 3
    for kat in sorted(per_kategori.keys()):
        items = per_kategori[kat]
        subtotal = sum(i["Harga"] for i in items)
        persen = (subtotal / total_keseluruhan * 100) if total_keseluruhan > 0 else 0
        warna = WARNA_KATEGORI.get(kat, "EEEEEE")

        cells_data = [kat, len(items), subtotal, f"{persen:.1f}%"]
        for col, val in enumerate(cells_data, 1):
            c = ws_rekap.cell(row=row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=warna)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left")
            if col == 3:
                c.number_format = '#,##0'
        row += 1

    # Baris total
    total_row = row
    ws_rekap.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws_rekap.cell(total_row, 2, len(data)).font = Font(bold=True)
    total_cell = ws_rekap.cell(total_row, 3, total_keseluruhan)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0'
    ws_rekap.cell(total_row, 4, "100%").font = Font(bold=True)
    for col in range(1, 5):
        ws_rekap.cell(total_row, col).fill = PatternFill("solid", fgColor="ECF0F1")

    # Lebar kolom
    ws_rekap.column_dimensions["A"].width = 20
    ws_rekap.column_dimensions["B"].width = 18
    ws_rekap.column_dimensions["C"].width = 18
    ws_rekap.column_dimensions["D"].width = 15

    # ── Sheet per KATEGORI ────────────────────────────────────────────────────
    for kat, items in sorted(per_kategori.items()):
        ws = wb.create_sheet(title=kat[:31])  # Excel max 31 char
        warna = WARNA_KATEGORI.get(kat, "EEEEEE")

        # Judul
        ws.merge_cells("A1:D1")
        judul_kat = ws["A1"]
        judul_kat.value = f"{kat} — {label_bulan}"
        judul_kat.font = Font(bold=True, size=12, color="FFFFFF")
        judul_kat.fill = PatternFill("solid", fgColor="2C3E50")
        judul_kat.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 25

        # Header
        hdrs = ["Tanggal", "Jam", "Nama Item", "Harga (Rp)"]
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(2, col, h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=warna)
            c.alignment = Alignment(horizontal="center")

        # Data
        for i, item in enumerate(sorted(items, key=lambda x: (x["Tanggal"], x["Jam"])), 3):
            ws.cell(i, 1, item["Tanggal"])
            ws.cell(i, 2, item["Jam"])
            ws.cell(i, 3, item["Nama Item"])
            harga_cell = ws.cell(i, 4, item["Harga"])
            harga_cell.number_format = '#,##0'
            if i % 2 == 0:
                for col in range(1, 5):
                    ws.cell(i, col).fill = PatternFill("solid", fgColor="F8F9FA")

        # Total baris
        total_baris = len(items) + 3
        ws.cell(total_baris, 3, "TOTAL").font = Font(bold=True)
        total_kat = ws.cell(total_baris, 4, sum(i["Harga"] for i in items))
        total_kat.font = Font(bold=True)
        total_kat.number_format = '#,##0'

        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 15

    # Simpan file
    os.makedirs("/tmp/rekap", exist_ok=True)
    filepath = f"/tmp/rekap/{nama_file}"
    wb.save(filepath)
    return filepath


async def buat_excel_rekap() -> str | None:
    """Export Excel bulan ini"""
    now = datetime.now()
    bulan_str = now.strftime("%Y-%m")
    label_bulan = now.strftime("%B %Y")
    nama_file = f"Rekap_{bulan_str}.xlsx"

    data = get_data_dari_sheets(bulan_str)
    return buat_excel_dari_data(data, nama_file, label_bulan)


async def buat_excel_rekap_bulan_lalu(bulan: int, tahun: int) -> str | None:
    """Export Excel bulan lalu (dipanggil scheduler tanggal 1)"""
    bulan_str = f"{tahun:04d}-{bulan:02d}"
    label_bulan = datetime(tahun, bulan, 1).strftime("%B %Y")
    nama_file = f"Rekap_{bulan_str}.xlsx"

    data = get_data_dari_sheets(bulan_str)
    return buat_excel_dari_data(data, nama_file, label_bulan)
